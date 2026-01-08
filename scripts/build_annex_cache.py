"""Convert annexes XLSX into JSON caches for rules/codelists.

Usage:
    python scripts/build_annex_cache.py --src ../specifications-externes-v3.1/2-\ Annexes_v3.1 --out ../MCP/data/annexes_cache

Notes:
- Requires openpyxl.
- Currently extracts sheet headers and rows as-is; tailor extractors per annex.
"""

import argparse
import json
from pathlib import Path
from typing import Any
import openpyxl
from datetime import datetime, date


def serialize_cell(val: Any) -> Any:
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return val


def load_workbook(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([serialize_cell(cell) for cell in row])
        data[sheet] = rows
    return data


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", type=Path, required=True, help="Directory containing annex XLSX")
    parser.add_argument("--out", type=Path, required=True, help="Output directory for JSON cache")
    args = parser.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    for xlsx in args.src.glob("*.xlsx"):
        content = load_workbook(xlsx)
        out_path = args.out / f"{xlsx.stem}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(content, f, ensure_ascii=False, indent=2)
        print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
